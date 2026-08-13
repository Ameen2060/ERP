"""Excel (.xlsx), CSV and PDF export for the reports. Each report is flattened to a
(title, headers, rows) table pulled live from the report services, so an export always
reflects current data. One generic renderer per format serves every report."""

from __future__ import annotations

import csv
import io
import json
from datetime import date
from decimal import Decimal

from fastapi import HTTPException
from fastapi.responses import Response, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from .schemas import CTComputation
from .services import advances, budgets, inventory, ledger, projects, reports
from .services.ledger import LedgerError

_CT_STATUS_LABEL = {
    "draft": "DRAFT",
    "provisional": "PROVISIONAL — AWAITING SME REVIEW",
    "sme_reviewed": "SME REVIEWED — AWAITING VALIDATION",
    "validated": "VALIDATED — FILING-READY",
    "rejected": "REJECTED — REQUIRES REWORK",
}


def _m(v) -> str:
    return f"{Decimal(str(v)):,.2f}"


def _aging_table(report, party_label: str, with_risk: bool):
    headers = [party_label, "Current", "1-30", "31-60", "61-90", "91-120", "120+", "Total", "Overdue"]
    if with_risk:
        headers.append("Risk")
    rows = []
    for r in report.rows:
        name = getattr(r, "customer_name", None) or getattr(r, "vendor_name", "")
        row = [name, r.current, r.d1_30, r.d31_60, r.d61_90, r.d91_120, r.d120_plus, r.total, r.overdue]
        if with_risk:
            row.append(getattr(r, "risk", ""))
        rows.append(row)
    total = ["TOTAL", report.current, report.d1_30, report.d31_60, report.d61_90,
             report.d91_120, report.d120_plus, report.total, report.total_overdue]
    if with_risk:
        total.append("")
    rows.append(total)
    return rows, headers


def _doc_table(report):
    headers = ["Party", "TRN", "Number", "Ref", "Date", "Due", "Net", "VAT", "Gross", "Paid", "Outstanding", "Status"]
    rows = [
        [r.party_name, r.party_trn or "", r.number, r.ref or "", str(r.date), str(r.due_date or ""),
         r.net, r.vat, r.gross, r.paid, r.outstanding, r.status]
        for r in report.rows
    ]
    rows.append(["TOTAL", "", "", "", "", "", report.total_net, report.total_vat, report.total_gross,
                 report.total_paid, report.total_outstanding, ""])
    return rows, headers


def _statement_rows(section, out):
    out.append([section.title.upper(), ""])
    for ln in section.lines:
        out.append([f"  {ln.code}  {ln.name}", ln.amount])
    out.append([f"Total {section.title}", section.total])


def build_table(db: Session, report: str, params: dict):
    """Return (title, headers, rows) for the requested report."""
    start, end, as_of = params.get("start"), params.get("end"), params.get("as_of")
    group_by = params.get("group_by")

    if report == "ap-aging":
        return "Accounts Payable Aging", *reversed(_aging_table(reports.ap_aging(db, as_of=as_of), "Vendor", False))
    if report == "ar-aging":
        return "Accounts Receivable Aging", *reversed(_aging_table(reports.ar_aging(db, as_of=as_of), "Customer", True))
    if report == "invoice-by-vendor":
        return "Invoice by Vendor", *reversed(_doc_table(reports.invoice_by_vendor(db, group_by=group_by or "vendor", start=start, end=end)))
    if report == "sales-by-customer":
        return "Sales by Customer", *reversed(_doc_table(reports.sales_by_customer(db, group_by=group_by or "customer", start=start, end=end)))
    if report == "asset-register":
        rep = reports.fixed_asset_register(
            db, category=params.get("category"), location=params.get("location"),
            department=params.get("department"), project=params.get("project"), status=params.get("status"))
        headers = ["Asset ID", "Asset", "Category", "Purchase Date", "Total Cost",
                   "Accumulated Depreciation", "Net Book Value", "Status"]
        rows = [[r.asset_code, r.name, r.category or "", str(r.purchase_date), r.total_cost,
                 r.accumulated_depreciation, r.net_book_value, r.status] for r in rep.rows]
        rows.append(["", "TOTAL", "", "", rep.total_cost, rep.total_accumulated_depreciation, rep.total_net_book_value, ""])
        return "Fixed Asset Register", headers, rows
    if report == "chart-of-accounts":
        accts = ledger.list_accounts(db)
        headers = ["Code", "Account", "Type", "Parent", "Normal Balance", "Group", "Active"]
        rows = [[a.code, a.name, a.type.value, a.parent_code or "", a.normal_balance.value,
                 "Yes" if a.is_group else "", "Yes" if a.is_active else "No"] for a in accts]
        return "Chart of Accounts", headers, rows
    if report == "trial-balance":
        tb = ledger.trial_balance(db, as_of=as_of)
        headers = ["Code", "Account", "Type", "Debit", "Credit"]
        order = [("asset", "ASSETS"), ("liability", "LIABILITIES"), ("equity", "EQUITY"),
                 ("income", "INCOME"), ("expense", "EXPENSES")]
        rows: list = []
        for key, label in order:
            grp = [r for r in tb.rows if r.type.value == key]
            if not grp:
                continue
            rows.append([label, "", "", "", ""])
            sd = sc = Decimal(0)
            for r in grp:
                sd += r.debit
                sc += r.credit
                rows.append([r.code, r.name, r.type.value, r.debit, r.credit])
            rows.append(["", f"Subtotal — {label.title()}", "", sd, sc])
        rows.append(["", "TOTAL", "", tb.total_debit, tb.total_credit])
        return "Trial Balance", headers, rows
    if report == "general-ledger":
        aid = params.get("account_id")
        if not aid:
            raise HTTPException(status_code=400, detail="general-ledger export requires an account_id.")
        try:
            gl = ledger.general_ledger(db, aid, start=start, end=end)
        except LedgerError as e:
            raise HTTPException(status_code=400, detail=str(e)) from e
        headers = ["#", "Date", "Memo", "Reference", "Source", "Debit", "Credit", "Balance"]
        rows = [["", "", "Opening balance", "", "", "", "", gl.opening_balance]]
        for r in gl.rows:
            rows.append([r.entry_no, str(r.date), r.memo or "", r.reference or "", r.source, r.debit, r.credit, r.balance])
        rows.append(["", "", "Closing balance", "", "", "", "", gl.closing_balance])
        return f"General Ledger — {gl.code} {gl.name}", headers, rows
    if report == "ct-computation":
        ct = reports.ct_computation(db, start, end)
        headers = ["Line", "Amount (AED)"]
        rows = [[ln.label, ln.amount] for ln in ct.lines]
        rows.append(["Effective rate", ct.effective_rate])
        return "Corporate Tax Computation (PROVISIONAL)", headers, rows
    if report in ("product-sales", "product-purchases"):
        rep = reports.product_movement(db, "issue" if report == "product-sales" else "receipt", start, end)
        headers = ["SKU", "Product", "Category", "Quantity", "Value (AED)", "Movements"]
        rows = [[r.sku, r.name, r.category or "", r.quantity, r.value, r.movements] for r in rep.rows]
        rows.append(["", "TOTAL", "", rep.total_quantity, rep.total_value, ""])
        return rep.title, headers, rows
    if report == "inventory-valuation":
        val = inventory.valuation(db)
        headers = ["SKU", "Product", "Category", "Method", "On hand", "Avg cost", "Value (AED)"]
        rows = [[r.sku, r.name, r.category or "", r.cost_method, r.on_hand, r.avg_cost, r.stock_value] for r in val.rows]
        rows.append(["", "TOTAL", "", "", "", "", val.total_value])
        return "Inventory Valuation", headers, rows
    if report in ("budget-vs-actual", "budget-forecast"):
        bid = params.get("budget_id")
        if not bid:
            raise HTTPException(status_code=400, detail=f"{report} export requires a budget_id.")
        if report == "budget-vs-actual":
            d = budgets.budget_vs_actual(db, bid)
            headers = ["Account", "Name", "Budget", "Actual", "Variance", "Variance %", "Remaining", "Utilization %"]
            rows = [[r["account_code"], r["account_name"], r["budget"], r["actual"], r["variance"],
                     r["variance_pct"], r["remaining"], r["utilization"]] for r in d["rows"]]
            rows.append(["", "TOTAL", d["total_budget"], d["total_actual"], d["total_variance"],
                         "", d["total_remaining"], d["utilization"]])
            return f"Budget vs Actual — {d['name']} (FY{d['fiscal_year']})", headers, rows
        d = budgets.forecast(db, bid)
        headers = ["Account", "Name", "Budget", "Actual YTD", "Projected (year-end)", "Projected variance", "Projected %"]
        rows = [[r["account_code"], r["account_name"], r["budget"], r["actual"], r["projected"],
                 r["projected_variance"], r["projected_pct"]] for r in d["rows"]]
        rows.append(["", "TOTAL", d["total_budget"], d["total_actual_ytd"], d["total_projected"],
                     d["projected_variance"], d["projected_utilization"]])
        return f"Budget Forecast — {d['name']} (FY{d['fiscal_year']}, {d['months_elapsed']}m elapsed)", headers, rows
    if report == "projects":
        d = projects.portfolio_report(db, status=params.get("status"))
        headers = ["Code", "Name", "Status", "Contract value", "Revenue", "Cost", "Gross profit",
                   "Margin %", "Receivables", "Payables", "Budget variance"]
        rows = [[r["code"], r["name"], r["status"], r["contract_value"], r["revenue"], r["cost"],
                 r["gross_profit"], r["gross_margin"], r["receivables"], r["payables"], r["budget_variance"]] for r in d["rows"]]
        t = d["totals"]
        rows.append(["", "TOTAL", "", t["contract_value"], t["revenue"], t["cost"], t["gross_profit"],
                     d["overall_margin"], t["receivables"], t["payables"], ""])
        return "Project Portfolio — P&L", headers, rows
    if report in ("advance-applications", "advance-aging"):
        side = params.get("side") or "customer"
        if report == "advance-applications":
            d = advances.advance_application_details(db, side=side)
            headers = ["Advance", side.capitalize(), "Date", "Applied to", "Amount (AED)"]
            rows = [[r["advance_number"], r["party"] or "", r["date"], r["target_number"], r["amount"]] for r in d["rows"]]
            rows.append(["", "", "", "TOTAL", d["total"]])
            return f"{side.capitalize()} Advance Applications", headers, rows
        d = advances.advance_aging(db, side=side)
        headers = ["Number", side.capitalize(), "Date", "Age (days)", "Bucket", "Outstanding (AED)"]
        rows = [[r["number"], r.get("customer_name") or r.get("vendor_name") or "", r["date"],
                 r.get("age_days", ""), r.get("bucket", ""), r["available"]] for r in d["rows"]]
        rows.append(["", "", "", "", "TOTAL", d["total_outstanding"]])
        return f"{side.capitalize()} Advance Aging", headers, rows
    if report in ("customer-statement", "vendor-statement"):
        is_cust = report == "customer-statement"
        pid = params.get("customer_id") if is_cust else params.get("vendor_id")
        if not pid:
            raise HTTPException(status_code=400, detail=f"{report} export requires a {'customer' if is_cust else 'vendor'}_id.")
        st = reports.customer_statement(db, pid, start, end) if is_cust else reports.vendor_statement(db, pid, start, end)
        headers = ["Date", "Type", "Reference", "Debit", "Credit", "Balance"]
        rows: list = [["", "Opening balance", "", "", "", st.opening_balance]]
        for ln in st.lines:
            rows.append([str(ln.date), ln.type, ln.reference or "", ln.debit, ln.credit, ln.balance])
        rows.append(["", "Closing balance", "", st.total_debit, st.total_credit, st.closing_balance])
        return f"{'Customer' if is_cust else 'Vendor'} Statement — {st.party_name}", headers, rows
    if report == "vat-return":
        rep = reports.vat_return(db, start=start, end=end)
        headers = ["Box", "Description", "Amount (AED)", "VAT (AED)"]
        rows = [[b.box, b.label, b.amount, b.vat] for b in rep.boxes]
        rows.append(["", "NET VAT DUE" + (" (refundable)" if rep.is_refund else " (payable)"), "", rep.net_vat_due])
        return "VAT Return (FTA VAT201)", headers, rows
    if report == "income-statement":
        pl = reports.income_statement(db, start=start, end=end)
        rows: list = []
        _statement_rows(pl.revenue, rows)
        _statement_rows(pl.cost_of_sales, rows)
        rows.append(["GROSS PROFIT", pl.gross_profit])
        _statement_rows(pl.operating_expenses, rows)
        rows.append(["NET PROFIT", pl.net_profit])
        return "Profit & Loss", ["Line", "Amount"], rows
    if report == "balance-sheet":
        bs = reports.balance_sheet(db, as_of=as_of or end)
        rows = []
        _statement_rows(bs.assets, rows)
        rows.append(["TOTAL ASSETS", bs.total_assets])
        _statement_rows(bs.liabilities, rows)
        rows.append(["Current period earnings", bs.current_earnings])
        _statement_rows(bs.equity, rows)
        rows.append(["TOTAL EQUITY", bs.total_equity])
        rows.append(["TOTAL LIABILITIES + EQUITY", Decimal(bs.total_liabilities) + Decimal(bs.total_equity)])
        return "Balance Sheet", ["Line", "Amount"], rows
    if report == "cash-flow":
        cf = reports.cash_flow(db, start=start, end=end)
        rows = [["Opening cash & bank", cf.opening_cash]]
        for ln in cf.lines:
            rows.append([f"  {ln.name}", ln.amount])
        rows.append(["Net change", cf.net_change])
        rows.append(["Closing cash & bank", cf.closing_cash])
        return "Cash Flow", ["Line", "Amount"], rows
    raise HTTPException(status_code=404, detail=f"Unknown report '{report}'.")


def _native(v):
    return float(v) if isinstance(v, Decimal) else v


def _is_num(v) -> bool:
    return isinstance(v, Decimal) or isinstance(v, (int, float))


def _company_line(db: Session) -> str:
    from .services import organization
    h = organization.company_header(db)
    parts = [h["name"]] if h.get("name") else []
    if h.get("trn"):
        parts.append("TRN " + h["trn"])
    return " · ".join(parts)


_FILTER_LABELS = {
    "start": "From", "end": "To", "as_of": "As of", "group_by": "Grouped by",
    "category": "Category", "location": "Location", "department": "Department",
    "project": "Project", "status": "Status", "account_id": "Account",
    "customer_id": "Customer", "vendor_id": "Vendor", "budget_id": "Budget", "side": "Side",
}


def _filter_meta(params: dict) -> list:
    """Human-readable 'filters used' rows for the PDF header (only non-empty params)."""
    return [(_FILTER_LABELS[k], str(v)) for k, v in params.items()
            if k in _FILTER_LABELS and v not in (None, "", "None")]


def export_response(db: Session, report: str, fmt: str, params: dict) -> Response:
    title, headers, rows = build_table(db, report, params)
    stamp = str(params.get("as_of") or params.get("end") or date.today())
    fname = f"{report}-{stamp}"

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow([f"{title} — generated {date.today()}"])
        w.writerow(headers)
        for r in rows:
            w.writerow([_native(c) for c in r])
        return Response(content=buf.getvalue(), media_type="text/csv",
                        headers={"Content-Disposition": f'attachment; filename="{fname}.csv"'})

    if fmt == "pdf":
        meta = _filter_meta(params) or None
        return _pdf(title, headers, rows, fname, company=_company_line(db), meta_rows=meta)

    # xlsx — professional formatting: title/company, frozen + auto-filtered header, number
    # formats on numeric columns, auto-fit widths.
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = report[:31]
    ws.append([title])
    ws["A1"].font = Font(bold=True, size=13)
    company = _company_line(db)
    if company:
        ws.append([company])
    ws.append([f"Generated {date.today().isoformat()}  ·  Currency AED"])
    for k, v in _filter_meta(params):
        ws.append([f"{k}: {v}"])
    ws.append([])
    header_row = ws.max_row + 1
    ws.append(headers)
    hfill = PatternFill("solid", fgColor="2563EB")
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    numeric_cols = {i for i, h in enumerate(headers)
                    if any(k in h.lower() for k in ("amount", "debit", "credit", "vat", "total", "net",
                            "gross", "paid", "outstanding", "value", "cost", "balance", "budget",
                            "revenue", "profit", "receivable", "payable", "actual", "projected",
                            "variance", "margin", "%", "contract", "current", "overdue", "1-30",
                            "31-60", "61-90", "91-120", "120+", "qty", "quantity", "on hand"))}
    for r in rows:
        ws.append([_native(c) for c in r])
        for i in numeric_cols:
            cell = ws.cell(row=ws.max_row, column=i + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
    last = ws.max_row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)          # freeze the header row
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last}"  # header filters
    for idx in range(1, len(headers) + 1):
        col = get_column_letter(idx)
        width = max((len(str(ws.cell(row=r, column=idx).value)) for r in range(header_row, last + 1)
                     if ws.cell(row=r, column=idx).value is not None), default=10)
        ws.column_dimensions[col].width = min(max(width + 2, 12), 48)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})


# ── Per-transaction Excel (reusable across document kinds) ───────────────────────────────
def _document_data(db: Session, kind: str, doc_id: str):
    """Normalise a transaction into (title, meta[(k,v)], line_headers, line_rows, totals[(k,v)])."""
    from .services import credit_notes, expenses as exp_svc, ledger as ledger_svc, purchases, sales
    if kind == "invoice":
        d = sales.get_invoice(db, doc_id)
        meta = [("Invoice #", d.number), ("Date", str(d.date)), ("Due", str(d.due_date or "—")),
                ("Customer", d.customer_name), ("Currency", d.currency), ("Status", d.status)]
        lh = ["Description", "Qty", "Unit price", "Net", "VAT", "Total"]
        lr = [[l.description, _native(l.quantity), _native(l.unit_price), _native(l.net_amount),
               _native(l.vat_amount), _native(l.line_total)] for l in d.lines]
        tot = [("Net", _native(d.net_total)), ("VAT", _native(d.vat_total)), ("Total", _native(d.grand_total)),
               ("Paid", _native(d.amount_paid)), ("Balance due", _native(d.balance_due))]
        return f"Tax Invoice {d.number}", meta, lh, lr, tot
    if kind == "bill":
        d = purchases.get_bill(db, doc_id)
        meta = [("Bill #", d.number), ("Date", str(d.date)), ("Vendor", d.vendor_name),
                ("Vendor ref", d.vendor_ref or "—"), ("Currency", d.currency), ("Status", d.status)]
        lh = ["Description", "Qty", "Unit price", "Net", "VAT", "Total"]
        lr = [[l.description, _native(l.quantity), _native(l.unit_price), _native(l.net_amount),
               _native(l.vat_amount), _native(l.line_total)] for l in d.lines]
        tot = [("Net", _native(d.net_total)), ("VAT", _native(d.vat_total)), ("Total", _native(d.grand_total)),
               ("Paid", _native(d.amount_paid)), ("Balance due", _native(d.balance_due))]
        return f"Vendor Bill {d.number}", meta, lh, lr, tot
    if kind in ("customer_cn", "vendor_cn"):
        d = credit_notes.get_customer_cn(db, doc_id) if kind == "customer_cn" else credit_notes.get_vendor_cn(db, doc_id)
        party = d.get("customer_name") or d.get("vendor_name")
        meta = [("Credit note #", d["number"]), ("Date", d["date"]), ("Party", party),
                ("Reason", d.get("reason") or "—"), ("Currency", d["currency"])]
        lh = ["Description", "Net", "VAT", "Total"]
        lr = [[l.get("description"), float(l["net_amount"]), float(l["vat_amount"]), float(l["line_total"])]
              for l in d["lines"]]
        tot = [("Net", float(d["net_total"])), ("VAT", float(d["vat_total"])), ("Total credit", float(d["grand_total"]))]
        return f"Credit Note {d['number']}", meta, lh, lr, tot
    if kind == "expense":
        d = exp_svc.get_expense(db, doc_id)
        meta = [("Expense #", d["number"]), ("Date", d["date"]), ("Vendor/Payee", d.get("vendor_name") or d.get("payee_name") or "—"),
                ("Category", d.get("category") or "—"), ("Account", (d.get("expense_account_code") or "") + " " + (d.get("expense_account_name") or ""))]
        lh = ["Description", "Net", "VAT", "Total"]
        lr = [[d.get("description") or d.get("category") or "Expense", float(d["net_amount"]), float(d["vat_amount"]), float(d["total_amount"])]]
        tot = [("Net", float(d["net_amount"])), ("VAT", float(d["vat_amount"])), ("Total", float(d["total_amount"]))]
        return f"Expense {d['number']}", meta, lh, lr, tot
    if kind in ("customer_advance", "vendor_advance"):
        from .services import advances as adv_svc
        side = "customer" if kind == "customer_advance" else "vendor"
        d = adv_svc.get_customer_advance(db, doc_id) if side == "customer" else adv_svc.get_vendor_advance(db, doc_id)
        party = d.get("customer_name") or d.get("vendor_name")
        meta = [("Advance #", d["number"]), ("Date", d["date"]), (side.capitalize(), party),
                ("Reference", d.get("reference") or "—"), ("Currency", d["currency"]),
                ("VAT on advance", ("Yes @ " + str(d["vat_rate"])) if d.get("vat_applicable") else "No")]
        lh = ["Description", "Net", "VAT", "Amount"]
        lr = [[f"Advance {d['number']}", float(d["net_amount"]), float(d["vat_amount"]), float(d["amount"])]]
        tot = [("Amount", float(d["amount"])), ("Applied", float(d["applied"])), ("Available", float(d["available"]))]
        return f"Advance {d['number']}", meta, lh, lr, tot
    if kind == "journal":
        d = ledger_svc.get_entry(db, doc_id)
        meta = [("Entry #", d.entry_no), ("Date", str(d.date)), ("Memo", d.memo or "—"),
                ("Reference", d.reference or "—"), ("Source", d.source), ("Status", d.status)]
        lh = ["Account", "Name", "Debit", "Credit"]
        lr = [[l.account_code, l.account_name, _native(l.debit), _native(l.credit)] for l in d.lines]
        tot = [("Total debits", _native(d.total)), ("Total credits", _native(d.total))]
        return f"Journal Entry #{d.entry_no}", meta, lh, lr, tot
    if kind == "einvoice":
        from .services import einvoicing as einv_svc
        ei = einv_svc.get(db, doc_id)
        d = einv_svc.detail(db, ei)
        payload = d.get("payload") or {}
        lines = payload.get("lines") or []
        meta = [("Document #", ei.doc_number or "—"), ("Type", ei.doc_type_code),
                ("Direction", ei.direction), ("Party", ei.party_name or "—"),
                ("Party TRN", ei.party_trn or "—"), ("Currency", ei.currency),
                ("E-invoice status", d["status_label"]),
                ("System validation", "PASSED" if ei.system_validation_passed else "NOT PASSED"),
                ("Regulatory compliance", "CONFIRMED" if ei.regulatory_confirmed else "NOT CONFIRMED"),
                ("Schema", f"{ei.schema_id or '—'} {ei.schema_version or ''}".strip()),
                ("Notice", einv_svc.PROVISIONAL_NOTICE if ei.provisional else "")]
        lh = ["Description", "Qty", "Unit price", "Net", "VAT rate", "VAT", "Total"]
        lr = [[l.get("description"), float(l.get("quantity") or 0), float(l.get("unit_price") or 0),
               float(l.get("net_amount") or 0), float(l.get("vat_rate") or 0),
               float(l.get("vat_amount") or 0), float(l.get("line_total") or 0)] for l in lines]
        tot = [("Net", _native(ei.net_total)), ("VAT", _native(ei.vat_total)),
               ("Gross", _native(ei.grand_total))]
        return f"E-Invoice {ei.doc_number or ei.id[:8]}", meta, lh, lr, tot
    raise HTTPException(status_code=400, detail=f"Excel not supported for '{kind}'.")


def document_excel(db: Session, kind: str, doc_id: str) -> StreamingResponse:
    """A professional single-transaction workbook: company header, meta block, line items and
    totals — reusable across invoice/bill/credit-note/expense/journal."""
    from openpyxl.styles import Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    title, meta, lh, lr, tot = _document_data(db, kind, doc_id)
    wb = Workbook()
    ws = wb.active
    ws.title = kind[:31]
    ws.append([title]); ws["A1"].font = Font(bold=True, size=13)
    company = _company_line(db)
    if company:
        ws.append([company])
    ws.append([f"Generated {date.today().isoformat()}"])
    ws.append([])
    for k, v in meta:
        ws.append([k, v]); ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.append([])
    hrow = ws.max_row + 1
    ws.append(lh)
    hfill = PatternFill("solid", fgColor="2563EB")
    for c in ws[hrow]:
        c.font = Font(bold=True, color="FFFFFF"); c.fill = hfill
    numcols = {i for i, h in enumerate(lh) if any(k in h.lower() for k in
               ("qty", "price", "net", "vat", "total", "debit", "credit"))}
    for r in lr:
        ws.append(r)
        for i in numcols:
            cell = ws.cell(row=ws.max_row, column=i + 1)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"; cell.alignment = Alignment(horizontal="right")
    ws.append([])
    for k, v in tot:
        ws.append([""] * (len(lh) - 2) + [k, v] if len(lh) > 2 else [k, v])
        rc = ws.cell(row=ws.max_row, column=len(lh) - 1 if len(lh) > 2 else 1)
        rc.font = Font(bold=True)
        vc = ws.cell(row=ws.max_row, column=len(lh) if len(lh) > 2 else 2)
        if isinstance(vc.value, (int, float)):
            vc.number_format = "#,##0.00"; vc.alignment = Alignment(horizontal="right"); vc.font = Font(bold=True)
    for idx in range(1, len(lh) + 1):
        col = get_column_letter(idx)
        width = max((len(str(ws.cell(row=r, column=idx).value)) for r in range(hrow, ws.max_row + 1)
                     if ws.cell(row=r, column=idx).value is not None), default=12)
        ws.column_dimensions[col].width = min(max(width + 2, 12), 48)
    out = io.BytesIO(); wb.save(out); out.seek(0)
    num = meta[0][1] if meta else doc_id
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{kind}-{num}.xlsx"'})


def ct_review_export(review, fmt: str, final: bool = False) -> Response:
    """Export a CT review's snapshotted computation, always carrying the provisional /
    validation control. Unless the review is validated, the document is stamped
    'PROVISIONAL — REQUIRES SME VALIDATION'; a validated review shows the SME sign-off."""
    comp = CTComputation.model_validate(json.loads(review.snapshot_json or "{}"))
    validated = review.status == "validated"
    period = f"{review.period_start or '—'} to {review.period_end or '—'}"
    stamp = str(review.period_end or date.today())
    fname = f"corporate-tax-{'final' if (final and validated) else 'provisional'}-{stamp}"

    # Control / status block shown at the top of every export.
    if validated:
        disclaimer = (
            f"SME-VALIDATED on {str(review.validated_at)[:10]} by {review.sme_name or '—'}. "
            "This computation has been reviewed and validated by a qualified tax specialist. "
            "Corporate Tax figures remain the taxpayer's responsibility.")
    else:
        disclaimer = (
            "PROVISIONAL — REQUIRES SME VALIDATION. This Corporate Tax computation has NOT been "
            "validated by a qualified tax specialist and MUST NOT be used for filing. Accounting "
            "treatment is not assumed to equal tax treatment.")

    meta_rows = [
        ["Status", _CT_STATUS_LABEL.get(review.status, review.status)],
        ["Period", period],
        ["Prepared by", review.prepared_by or "—"],
        ["SME reviewer", review.sme_name or "—"],
        ["Validated", "Yes" if validated else "No"],
        ["Legal basis", comp.legal_ref],
    ]
    comp_rows = [[ln.label, ln.amount] for ln in comp.lines]
    comp_rows.append(["Effective rate", comp.effective_rate])
    title = "Corporate Tax Computation — " + ("FINAL (Validated)" if (final and validated) else "PROVISIONAL")

    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.title = "CT Computation"
        ws.append([title]); ws["A1"].font = Font(bold=True, size=13)
        ws.append([disclaimer]); ws["A2"].font = Font(bold=True, color=("1a7f37" if validated else "b91c1c"))
        ws.append([])
        for r in meta_rows:
            ws.append(r); ws[f"A{ws.max_row}"].font = Font(bold=True)
        ws.append([])
        ws.append(["Line", "Amount (AED)"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
        for r in comp_rows:
            ws.append([_native(c) for c in r])
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 14), 70)
        out = io.BytesIO(); wb.save(out); out.seek(0)
        return StreamingResponse(
            out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}.xlsx"'})

    # pdf
    return _pdf("Corporate Tax Computation", ["Line", "Amount (AED)"], comp_rows, fname,
                disclaimer=disclaimer, disclaimer_ok=validated, meta_rows=meta_rows)


def _pdf(title: str, headers: list, rows: list, fname: str,
         disclaimer: str | None = None, disclaimer_ok: bool = False, meta_rows: list | None = None,
         company: str | None = None) -> Response:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    wide = len(headers) > 5
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4) if wide else A4,
                            leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    elems = []
    if company:
        elems.append(Paragraph(f"<b>{company}</b>", styles["Normal"]))
    elems += [Paragraph(title, styles["Title"]),
              Paragraph(f"Generated {date.today().isoformat()} · Currency AED", styles["Normal"]),
              Spacer(1, 8)]
    if disclaimer:
        band = colors.HexColor("#dcfce7") if disclaimer_ok else colors.HexColor("#fee2e2")
        edge = colors.HexColor("#1a7f37") if disclaimer_ok else colors.HexColor("#b91c1c")
        dp = Table([[Paragraph(f"<b>{disclaimer}</b>", styles["Normal"])]], colWidths=["100%"])
        dp.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), band),
            ("BOX", (0, 0), (-1, -1), 1, edge),
            ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ]))
        elems += [dp, Spacer(1, 8)]
    if meta_rows:
        mt = Table([[Paragraph(f"<b>{k}</b>", styles["Normal"]), Paragraph(str(v), styles["Normal"])]
                    for k, v in meta_rows], colWidths=[120, 360])
        mt.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f4f6f9")),
        ]))
        elems += [mt, Spacer(1, 10)]
    numeric_cols = {i for i, h in enumerate(headers)
                    if any(k in h.lower() for k in ("amount", "debit", "credit", "vat", "total", "net",
                                                    "gross", "paid", "outstanding", "value", "cost", "balance",
                                                    "current", "overdue", "1-30", "31-60", "61-90", "91-120", "120+"))}
    data = [headers]
    for r in rows:
        data.append([_m(c) if (i in numeric_cols and _is_num(c)) else ("" if c is None else str(c)) for i, c in enumerate(r)])
    table = Table(data, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for i in numeric_cols:
        style.append(("ALIGN", (i, 0), (i, -1), "RIGHT"))
    table.setStyle(TableStyle(style))
    elems.append(table)

    from datetime import datetime, timezone
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#888888"))
        w, _h = doc_.pagesize
        canvas.drawString(14 * mm, 8 * mm, f"{title} · generated {ts}")
        canvas.drawRightString(w - 14 * mm, 8 * mm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(elems, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{fname}.pdf"'})
