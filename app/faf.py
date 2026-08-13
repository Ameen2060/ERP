"""FAF — FTA VAT Audit File (UAE).

Builds the audit workbook the FTA expects during a VAT audit: a Required-Information sheet,
a VAT Return summary, and one transaction-listing sheet per VAT201 box — all generated from
the posted sales invoices, vendor bills and the VAT return, nothing hardcoded.

Note: this is an FAF-*format* workbook produced from the accounting ledger. The official FTA
FAF template (with the full 12-question header) lives in the VAT Platform; here we reproduce
its structure and content from the general ledger.
"""

from __future__ import annotations

import io
from datetime import date, datetime, timezone
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from .config import get_settings
from .models import Customer, SalesInvoice, SalesInvoiceLine, Vendor, VendorBill, VendorBillLine
from .services import reports

_POSTED = ("posted", "partial", "paid")
_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="2563EB")
_TITLE = Font(bold=True, size=13)


def _num(v) -> float:
    return float(Decimal(str(v or 0)))


def _sheet(wb: Workbook, name: str, headers: list[str]):
    ws = wb.create_sheet(title=name[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = _HEAD
        c.fill = _HEAD_FILL
        c.alignment = Alignment(horizontal="center")
    return ws


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(width + 2, 12), 48)


def _sales_rows(db: Session, start, end, zero: bool):
    stmt = (
        select(SalesInvoiceLine, SalesInvoice, Customer)
        .join(SalesInvoice, SalesInvoiceLine.invoice_id == SalesInvoice.id)
        .join(Customer, SalesInvoice.customer_id == Customer.id, isouter=True)
        .where(SalesInvoice.status.in_(_POSTED))
        .order_by(SalesInvoice.date, SalesInvoice.number)
    )
    if start:
        stmt = stmt.where(SalesInvoice.date >= start)
    if end:
        stmt = stmt.where(SalesInvoice.date <= end)
    for line, inv, cust in db.execute(stmt).all():
        is_zero = Decimal(line.vat_rate) == 0
        if is_zero != zero:
            continue
        yield [str(inv.date), inv.number, cust.name if cust else "", (cust.trn if cust else "") or "",
               line.description or "", _num(line.net_amount), _num(line.vat_amount), _num(line.line_total)]


def _expense_rows(db: Session, start, end):
    stmt = (
        select(VendorBillLine, VendorBill, Vendor)
        .join(VendorBill, VendorBillLine.bill_id == VendorBill.id)
        .join(Vendor, VendorBill.vendor_id == Vendor.id, isouter=True)
        .where(VendorBill.status.in_(_POSTED))
        .order_by(VendorBill.date, VendorBill.number)
    )
    if start:
        stmt = stmt.where(VendorBill.date >= start)
    if end:
        stmt = stmt.where(VendorBill.date <= end)
    for line, bill, ven in db.execute(stmt).all():
        if Decimal(line.vat_rate) == 0:
            continue
        yield [str(bill.date), bill.number, bill.vendor_ref or "", ven.name if ven else "",
               (ven.trn if ven else "") or "", line.description or "",
               _num(line.net_amount), _num(line.vat_amount), _num(line.line_total)]


def build_faf(db: Session, start: date | None, end: date | None) -> bytes:
    settings = get_settings()
    from .services import organization
    company = organization.company_header(db)
    vat = reports.vat_return(db, start=start, end=end)
    wb = Workbook()

    # 1) Required information
    info = wb.active
    info.title = "Required Information"
    info["A1"] = "FTA VAT Audit File (FAF)"
    info["A1"].font = _TITLE
    rows = [
        ("Taxable person", company.get("name") or settings.app_name),
        ("TRN", company.get("trn") or ""),
        ("Period from", start.isoformat() if start else "All"),
        ("Period to", end.isoformat() if end else "All"),
        ("Currency", "AED"),
        ("Generated", datetime.now(timezone.utc).isoformat(timespec="seconds")),
        ("Net VAT due", _num(vat.net_vat_due)),
        ("Position", "Refundable" if vat.is_refund else "Payable"),
        ("Reconciled to GL", "Yes" if vat.reconciled else "No"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        info[f"A{i}"] = k
        info[f"A{i}"].font = Font(bold=True)
        info[f"B{i}"] = v
    _autosize(info)

    # 2) VAT Return summary (VAT201 boxes)
    ret = _sheet(wb, "VAT Return", ["Box", "Description", "Amount (AED)", "VAT (AED)"])
    for b in vat.boxes:
        ret.append([b.box, b.label, _num(b.amount), _num(b.vat)])
    _autosize(ret)

    # 3..) Transaction listings per box
    s1 = _sheet(wb, "Box 1 - Standard Sales",
                ["Date", "Invoice No", "Customer", "Customer TRN", "Description", "Net", "VAT", "Gross"])
    for r in _sales_rows(db, start, end, zero=False):
        s1.append(r)
    _autosize(s1)

    s4 = _sheet(wb, "Box 4 - Zero-Rated Sales",
                ["Date", "Invoice No", "Customer", "Customer TRN", "Description", "Net", "VAT", "Gross"])
    for r in _sales_rows(db, start, end, zero=True):
        s4.append(r)
    _autosize(s4)

    s9 = _sheet(wb, "Box 9 - Standard Expenses",
                ["Date", "Bill No", "Vendor Ref", "Vendor", "Vendor TRN", "Description", "Net", "VAT", "Gross"])
    for r in _expense_rows(db, start, end):
        s9.append(r)
    _autosize(s9)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
