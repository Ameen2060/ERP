"""Import engine: upload → read (xlsx/xls/csv/pdf) → validate → analyze → commit.

One generic pipeline serves three document kinds:
  * tb      — Trial Balance   → posts one balanced opening journal entry
  * report  — Financial report (Balance Sheet / P&L) → analysed + posts opening balances
  * gl      — General Ledger  → groups rows into balanced journal entries and posts them

Parsing produces a raw grid; a header row is detected and columns are mapped to semantic
fields by keyword. Nothing is template-specific.
"""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from sqlalchemy import select
from sqlalchemy.orm import Session

from .. import constants as C
from ..constants import AccountType, NormalBalance
from ..models import Account
from ..schemas import (
    ImportAnalysis,
    ImportCommitIn,
    ImportGroup,
    ImportResult,
    ImportRow,
    JournalEntryIn,
    JournalLineIn,
    TypeTotal,
    q,
)
from . import ledger

ZERO = Decimal(0)


class ImportError_(ValueError):
    """Domain error → HTTP 400."""


# ── File parsing → grid ─────────────────────────────────────────────────────────────────
def parse_grid(filename: str, data: bytes) -> list[list[str]]:
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "").lower()
    if ext == "csv":
        return _grid_csv(data)
    if ext == "xlsx":
        return _grid_xlsx(data)
    if ext == "xls":
        return _grid_xls(data)
    if ext == "pdf":
        return _grid_pdf(data)
    raise ImportError_(f"Unsupported file type '.{ext}'. Use .xlsx, .xls, .csv or .pdf.")


def _grid_csv(data: bytes) -> list[list[str]]:
    text = data.decode("utf-8-sig", errors="replace")
    return [[c.strip() for c in row] for row in csv.reader(io.StringIO(text)) if any(c.strip() for c in row)]


def _grid_xlsx(data: bytes) -> list[list[str]]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    grid = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            grid.append(cells)
    return grid


def _grid_xls(data: bytes) -> list[list[str]]:
    try:
        import xlrd
    except ImportError as e:  # pragma: no cover
        raise ImportError_("Reading .xls needs the 'xlrd' package — or save the file as .xlsx.") from e
    book = xlrd.open_workbook(file_contents=data)
    sheet = book.sheet_by_index(0)
    grid = []
    for r in range(sheet.nrows):
        cells = []
        for c in range(sheet.ncols):
            v = sheet.cell_value(r, c)
            cells.append("" if v == "" else (str(int(v)) if isinstance(v, float) and v.is_integer() else str(v)).strip())
        if any(cells):
            grid.append(cells)
    return grid


def _grid_pdf(data: bytes) -> list[list[str]]:
    import pdfplumber
    grid: list[list[str]] = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                for row in table:
                    cells = ["" if c is None else str(c).replace("\n", " ").strip() for c in row]
                    if any(cells):
                        grid.append(cells)
            if not tables:
                # Fallback: split text lines on runs of 2+ spaces.
                for line in (page.extract_text() or "").splitlines():
                    parts = [p.strip() for p in re.split(r"\s{2,}", line.strip()) if p.strip()]
                    if len(parts) >= 2:
                        grid.append(parts)
    if not grid:
        raise ImportError_("Could not extract any tabular data from the PDF.")
    return grid


# ── Header detection + column mapping ─────────────────────────────────────────────────────
_FIELD_KEYWORDS = {
    "code": ["account code", "acct code", "gl code", "code", "account no", "acc no", "a/c"],
    "name": ["account name", "account", "description", "particulars", "name", "account title", "gl name"],
    "debit": ["debit", "dr", "debit amount", "debit (aed)"],
    "credit": ["credit", "cr", "credit amount", "credit (aed)"],
    "balance": ["balance", "closing balance", "amount", "net", "balance (aed)"],
    "date": ["date", "posting date", "transaction date", "entry date", "trans date", "doc date"],
    "reference": ["reference", "ref", "voucher", "journal no", "journal", "entry no", "doc no", "document"],
    "memo": ["narration", "memo", "details", "remarks", "description"],
}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower())


def _detect_header(grid: list[list[str]]) -> int:
    """Row index whose cells best match known column keywords."""
    best_i, best_score = 0, -1
    flat_kw = {kw for kws in _FIELD_KEYWORDS.values() for kw in kws}
    for i, row in enumerate(grid[:15]):
        score = sum(1 for cell in row if _norm(cell) in flat_kw or any(k in _norm(cell) for k in ("debit", "credit", "account", "date", "balance")))
        if score > best_score:
            best_i, best_score = i, score
    return best_i if best_score > 0 else 0


def _map_columns(header: list[str]) -> dict[str, int]:
    norm = [_norm(h) for h in header]
    mapping: dict[str, int] = {}
    used: set[int] = set()
    for field, keywords in _FIELD_KEYWORDS.items():
        for kw in keywords:  # exact match first
            for j, h in enumerate(norm):
                if j not in used and h == kw:
                    mapping[field] = j
                    used.add(j)
                    break
            if field in mapping:
                break
        if field not in mapping:  # then contains
            for kw in keywords:
                for j, h in enumerate(norm):
                    if j not in used and h and kw in h:
                        mapping[field] = j
                        used.add(j)
                        break
                if field in mapping:
                    break
    return mapping


_NUM_RE = re.compile(r"-?\(?\d[\d,]*\.?\d*\)?")


def _num(s: str) -> Decimal:
    if s is None:
        return ZERO
    t = str(s).strip().replace(",", "")
    if not t or t in ("-", "—"):
        return ZERO
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()").replace("AED", "").replace("aed", "").strip()
    try:
        d = Decimal(t)
    except (InvalidOperation, ValueError):
        return ZERO
    return -d if neg else d


def _parse_date(s: str | None) -> date | None:
    if not s:
        return None
    s = str(s).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y", "%d %B %Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


_SUBTOTAL_RE = re.compile(r"\b(total|subtotal|sum|grand total|balance c/f|net (profit|loss|income))\b", re.I)


# ── Analysis ────────────────────────────────────────────────────────────────────────────
def _resolve_account(db: Session, code: str | None, name: str | None, accounts: list[Account]):
    # A supplied code is authoritative: match on it exactly, and if it doesn't exist treat
    # the row as a NEW account (never fall back to a fuzzy name match, which would wrongly
    # merge e.g. "Petty Cash Fund" into "Cash").
    if code:
        code = code.strip()
        for a in accounts:
            if a.code == code:
                return a
        return None
    if name:
        n = _norm(name)
        for a in accounts:
            if _norm(a.name) == n:
                return a
        for a in accounts:
            if n and (_norm(a.name) in n or n in _norm(a.name)):
                return a
    return None


def _infer_type(code: str | None) -> AccountType | None:
    if code and code[:1] in "12345":
        return {"1": AccountType.ASSET, "2": AccountType.LIABILITY, "3": AccountType.EQUITY,
                "4": AccountType.INCOME, "5": AccountType.EXPENSE}[code[0]]
    return None


def analyze(db: Session, kind: str, filename: str, data: bytes) -> ImportAnalysis:
    if kind not in ("tb", "gl", "report"):
        raise ImportError_("kind must be 'tb', 'gl' or 'report'.")
    grid = parse_grid(filename, data)
    if len(grid) < 2:
        raise ImportError_("File has no data rows.")
    hidx = _detect_header(grid)
    header = grid[hidx]
    mapping = _map_columns(header)
    warnings: list[str] = []
    accounts = list(db.execute(select(Account)).scalars())

    has_dr = "debit" in mapping
    has_cr = "credit" in mapping
    has_bal = "balance" in mapping
    if not (has_dr and has_cr) and not has_bal:
        warnings.append("No Debit/Credit or Balance columns detected — check the file layout.")
    if not has_dr and not has_cr and has_bal:
        warnings.append("Only a Balance column found: positive → debit, negative → credit.")

    def cell(row, field):
        j = mapping.get(field)
        return row[j] if j is not None and j < len(row) else ""

    rows: list[ImportRow] = []
    matched = new = unmatched = 0
    for r in range(hidx + 1, len(grid)):
        raw = grid[r]
        code = cell(raw, "code").strip() or None
        name = cell(raw, "name").strip() or None
        if not code and not name:
            continue
        # Skip subtotal / header lines in reports & TBs.
        if kind in ("tb", "report") and name and _SUBTOTAL_RE.search(name) and not code:
            continue
        debit = _num(cell(raw, "debit")) if has_dr else ZERO
        credit = _num(cell(raw, "credit")) if has_cr else ZERO
        if not has_dr and not has_cr and has_bal:
            bal = _num(cell(raw, "balance"))
            debit, credit = (bal, ZERO) if bal >= 0 else (ZERO, -bal)
        d = _parse_date(cell(raw, "date")) if "date" in mapping else None
        ref = cell(raw, "reference").strip() or None
        memo = cell(raw, "memo").strip() or None
        if debit == 0 and credit == 0 and kind != "gl":
            continue

        acct = _resolve_account(db, code, name, accounts)
        row = ImportRow(
            row_no=r, code=code, name=name, debit=q(debit), credit=q(credit), date=d,
            reference=ref, memo=memo,
        )
        if acct:
            row.matched_account_id = acct.id
            row.matched_code = acct.code
            row.matched_name = acct.name
            row.account_type = acct.type
            row.status = "ok" if not acct.is_group else "unmatched"
            if acct.is_group:
                row.note = "matched a group account (non-postable)"
                unmatched += 1
            else:
                matched += 1
        else:
            t = _infer_type(code)
            row.account_type = t.value if t else None
            row.status = "new"
            row.note = "no matching account — will be created on import" if t else "no match; provide a code to auto-create"
            new += 1
        rows.append(row)

    total_debit = q(sum((r.debit for r in rows), ZERO))
    total_credit = q(sum((r.credit for r in rows), ZERO))
    balanced = abs(total_debit - total_credit) < Decimal("0.01")

    groups: list[ImportGroup] = []
    by_type: list[TypeTotal] = []
    if kind == "gl":
        groups = _group_gl(rows, warnings)
    if kind == "report":
        by_type = _report_totals(rows)

    return ImportAnalysis(
        kind=kind, filename=filename, detected_columns=header, mapping={k: header[v] for k, v in mapping.items() if v < len(header)},
        rows=rows, total_debit=total_debit, total_credit=total_credit, balanced=balanced,
        matched_count=matched, new_count=new, unmatched_count=unmatched, groups=groups, by_type=by_type,
        warnings=warnings,
    )


def _group_gl(rows: list[ImportRow], warnings: list[str]) -> list[ImportGroup]:
    buckets: dict[str, list[ImportRow]] = {}
    for r in rows:
        key = r.reference or (r.date.isoformat() if r.date else "ungrouped")
        buckets.setdefault(key, []).append(r)
    if all((r.reference is None) for r in rows):
        warnings.append("No reference/voucher column — rows grouped by date. Add a reference column for precise grouping.")
    groups = []
    for key, rs in buckets.items():
        td = q(sum((r.debit for r in rs), ZERO))
        tc = q(sum((r.credit for r in rs), ZERO))
        groups.append(ImportGroup(
            reference=key, date=next((r.date for r in rs if r.date), None), total_debit=td,
            total_credit=tc, balanced=abs(td - tc) < Decimal("0.01"), row_nos=[r.row_no for r in rs],
        ))
    return groups


def _report_totals(rows: list[ImportRow]) -> list[TypeTotal]:
    agg: dict[str, list[Decimal]] = {}
    for r in rows:
        t = r.account_type or "unclassified"
        d, c = agg.setdefault(t, [ZERO, ZERO])
        agg[t] = [d + r.debit, c + r.credit]
    return [TypeTotal(type=t, debit=q(v[0]), credit=q(v[1]), net=q(v[0] - v[1])) for t, v in sorted(agg.items())]


# ── Commit ──────────────────────────────────────────────────────────────────────────────
def _ensure_account(db: Session, code: str | None, name: str | None, create_missing: bool, created: list[str]):
    accounts = list(db.execute(select(Account)).scalars())
    acct = _resolve_account(db, code, name, accounts)
    if acct and not acct.is_group:
        return acct
    if acct and acct.is_group:
        raise ImportError_(f"Account '{acct.code}' is a group account and cannot receive a balance.")
    if not create_missing:
        raise ImportError_(f"No matching account for '{code or name}'. Enable 'create missing' or add it to the Chart of Accounts.")
    t = _infer_type(code)
    if not t or not code:
        raise ImportError_(f"Cannot auto-create '{name or code}': need a numeric code (1xxx–5xxx) to infer the account type.")
    parent = db.execute(select(Account).where(Account.code == code[0] + "000")).scalar_one_or_none()
    from ..schemas import AccountIn
    out = ledger.create_account(db, AccountIn(
        code=code.strip(), name=(name or code).strip(), type=t,
        parent_code=parent.code if parent else None,
        normal_balance=NormalBalance.DEBIT if t in (AccountType.ASSET, AccountType.EXPENSE) else NormalBalance.CREDIT,
    ))
    created.append(out.code)
    return db.get(Account, out.id)


def commit(db: Session, data: ImportCommitIn) -> ImportResult:
    if data.kind not in ("tb", "gl", "report"):
        raise ImportError_("kind must be 'tb', 'gl' or 'report'.")
    created_accounts: list[str] = []
    skipped: list[str] = []

    if data.kind in ("tb", "report"):
        lines: list[JournalLineIn] = []
        imported = 0
        for row in data.rows:
            if q(row.debit) == 0 and q(row.credit) == 0:
                continue
            acct = _ensure_account(db, row.code, row.name, data.create_missing, created_accounts)
            lines.append(JournalLineIn(account_id=acct.id, debit=q(row.debit), credit=q(row.credit)))
            imported += 1
        if not lines:
            raise ImportError_("Nothing to import.")
        td = q(sum((ln.debit for ln in lines), ZERO))
        tc = q(sum((ln.credit for ln in lines), ZERO))
        if abs(td - tc) >= Decimal("0.01"):
            raise ImportError_(f"Trial balance does not balance: debits {td} vs credits {tc}.")
        entry = ledger.create_journal_entry(db, JournalEntryIn(
            date=data.opening_date or date.today(),
            memo=data.memo or ("Trial Balance import" if data.kind == "tb" else "Financial report import"),
            source="opening", lines=lines, auto_post=True,
        ))
        return ImportResult(
            kind=data.kind, journal_entry_ids=[entry.id], entries_created=1, accounts_created=created_accounts,
            rows_imported=imported, skipped=skipped,
            message=f"Imported {imported} balances as one opening journal entry (#{entry.entry_no}).",
        )

    # GL: group into entries and post each balanced group.
    buckets: dict[str, list] = {}
    for row in data.rows:
        key = row.reference or (row.date.isoformat() if row.date else "ungrouped")
        buckets.setdefault(key, []).append(row)
    entry_ids: list[str] = []
    imported = 0
    for key, rs in buckets.items():
        td = q(sum((r.debit for r in rs), ZERO))
        tc = q(sum((r.credit for r in rs), ZERO))
        if abs(td - tc) >= Decimal("0.01"):
            if data.skip_unbalanced:
                skipped.append(f"{key} (unbalanced {td}/{tc})")
                continue
            raise ImportError_(f"Group '{key}' does not balance: {td}/{tc}.")
        lines = []
        gdate = next((r.date for r in rs if r.date), None) or data.opening_date or date.today()
        for r in rs:
            if q(r.debit) == 0 and q(r.credit) == 0:
                continue
            acct = _ensure_account(db, r.code, r.name, data.create_missing, created_accounts)
            lines.append(JournalLineIn(account_id=acct.id, debit=q(r.debit), credit=q(r.credit),
                                       description=r.memo))
        if len(lines) < 2:
            skipped.append(f"{key} (fewer than 2 lines)")
            continue
        entry = ledger.create_journal_entry(db, JournalEntryIn(
            date=gdate, memo=data.memo or f"GL import {key}", reference=str(key)[:120],
            source="import", lines=lines, auto_post=True,
        ))
        entry_ids.append(entry.id)
        imported += len(lines)
    return ImportResult(
        kind="gl", journal_entry_ids=entry_ids, entries_created=len(entry_ids),
        accounts_created=created_accounts, rows_imported=imported, skipped=skipped,
        message=f"Imported {len(entry_ids)} journal entr{'y' if len(entry_ids) == 1 else 'ies'}"
                + (f"; skipped {len(skipped)}." if skipped else "."),
    )
