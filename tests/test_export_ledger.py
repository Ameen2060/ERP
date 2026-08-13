"""Export coverage for Chart of Accounts, Trial Balance (grouped by category), General
Ledger, and the FAF VAT audit file — in Excel, PDF and CSV."""

from __future__ import annotations

from fastapi.testclient import TestClient

from app.main import app


def _acct(client, code):
    return next(a["id"] for a in client.get("/api/accounts").json() if a["code"] == code)


def _seed_activity(client):
    cid = client.post("/api/sales/customers", json={"name": "Exp Cust", "trn": "100111222333001"}).json()["id"]
    client.post("/api/sales/invoices", json={
        "customer_id": cid, "date": "2029-01-05",
        "lines": [{"description": "Svc", "quantity": "1", "unit_price": "1000", "vat_rate": "0.05"}]})


def test_chart_of_accounts_export():
    with TestClient(app) as client:
        for fmt, magic in (("xlsx", b"PK"), ("pdf", b"%PDF-"), ("csv", b"Chart")):
            r = client.get("/api/export/chart-of-accounts", params={"format": fmt})
            assert r.status_code == 200, r.text
            assert r.content[: len(magic)] == magic


def test_trial_balance_grouped_export():
    with TestClient(app) as client:
        _seed_activity(client)
        csv = client.get("/api/export/trial-balance", params={"format": "csv"})
        assert csv.status_code == 200
        text = csv.text
        # Category group headers + subtotals present.
        assert "ASSETS" in text and "Subtotal — Assets" in text
        assert "TOTAL" in text
        assert client.get("/api/export/trial-balance", params={"format": "pdf"}).content[:5] == b"%PDF-"


def test_general_ledger_export_requires_account():
    with TestClient(app) as client:
        _seed_activity(client)
        # Missing account_id → 400.
        assert client.get("/api/export/general-ledger", params={"format": "xlsx"}).status_code == 400
        aid = _acct(client, "1100")  # Accounts Receivable has movement from the invoice
        r = client.get("/api/export/general-ledger", params={"format": "pdf", "account_id": aid})
        assert r.status_code == 200 and r.content[:5] == b"%PDF-"


def test_faf_vat_audit_file():
    with TestClient(app) as client:
        _seed_activity(client)
        r = client.get("/api/reports/faf", params={"start": "2029-01-01", "end": "2029-12-31"})
        assert r.status_code == 200
        assert "spreadsheetml" in r.headers["content-type"]
        assert r.content[:2] == b"PK"  # xlsx zip
        # Confirm expected sheets exist.
        import io
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        assert "Required Information" in wb.sheetnames
        assert "VAT Return" in wb.sheetnames
        assert any("Standard Sales" in s for s in wb.sheetnames)
